# ============================================================
# fintrack — Export router: /api/v1/export
# File: backend/app/routers/export.py
# Version: 1.0 — 2026-04-16
#
# Endpoints:
#   GET /api/v1/export/excel   download Excel workbook (4 sheets)
#   GET /api/v1/export/pdf     download PDF report
#
# Both endpoints require X-Fintrack-Password header for
# decrypting transaction descriptions.
# ============================================================

import io
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query, Header, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)

from app.database import get_db
from app.models.user import UserKey
from app.services.auth import CurrentUser
from app.services.encryption import decrypt, derive_key

logger = logging.getLogger("fintrack.export")
router = APIRouter()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_enc_key(user_id, password: str, db: Session) -> bytes:
    user_key = db.query(UserKey).filter(UserKey.user_id == user_id).first()
    if not user_key:
        raise HTTPException(status_code=500, detail="Key material not found")
    return derive_key(password, bytes(user_key.kdf_salt))


def _get_category_summary(user_id, year, db):
    params = {"uid": str(user_id)}
    yf = "AND t.year_num = :year" if year else ""
    if year: params["year"] = year
    rows = db.execute(text(f"""
        SELECT t.category_name,
               COALESCE(t.subcategory, t.category_name) AS subcategory,
               SUM(t.amount) AS total, COUNT(*) AS txn_count
        FROM transactions t
        WHERE t.user_id = :uid {yf}
        GROUP BY t.category_name, t.subcategory
        ORDER BY total DESC
    """), params).fetchall()
    grand = sum(float(r.total) for r in rows)
    return rows, grand


def _get_monthly_pivot(user_id, year, db):
    params = {"uid": str(user_id)}
    yf = "AND t.year_num = :year" if year else ""
    if year: params["year"] = year
    rows = db.execute(text(f"""
        SELECT t.year_num, t.month_num,
               TO_CHAR(TO_DATE(t.year_num::text||'-'||LPAD(t.month_num::text,2,'0')||'-01','YYYY-MM-DD'),'Mon YYYY') AS month_label,
               t.category_name,
               COALESCE(t.subcategory, t.category_name) AS subcategory,
               SUM(t.amount) AS total
        FROM transactions t
        WHERE t.user_id = :uid {yf}
        GROUP BY t.year_num, t.month_num, t.category_name, t.subcategory
        ORDER BY t.year_num, t.month_num, t.category_name
    """), params).fetchall()
    return rows


def _get_transactions(user_id, year, enc_key, db):
    params = {"uid": str(user_id)}
    yf = "AND t.year_num = :year" if year else ""
    if year: params["year"] = year
    rows = db.execute(text(f"""
        SELECT t.txn_date, t.amount, t.description,
               t.category_name,
               COALESCE(t.subcategory, t.category_name) AS subcategory,
               t.is_essential, a.provider
        FROM transactions t
        JOIN accounts a ON a.id = t.account_id
        WHERE t.user_id = :uid {yf}
        ORDER BY t.txn_date DESC
    """), params).fetchall()
    result = []
    for r in rows:
        try:
            desc = decrypt(r.description, enc_key)
        except Exception:
            desc = "***"
        result.append({
            "date":        r.txn_date.isoformat(),
            "description": desc,
            "amount":      float(r.amount),
            "category":    r.category_name,
            "subcategory": r.subcategory,
            "essential":   "Yes" if r.is_essential else "No",
            "provider":    r.provider,
        })
    return result


# ── Excel Export ──────────────────────────────────────────────────────────────

HEADER_FILL    = PatternFill("solid", fgColor="1565C0")
HEADER_FONT    = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL       = PatternFill("solid", fgColor="EEF2FF")
TOTAL_FONT     = Font(bold=True)
TOTAL_FILL     = PatternFill("solid", fgColor="DBEAFE")
BORDER_SIDE    = Side(style="thin", color="CBD5E1")
CELL_BORDER    = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE
)
MONEY_FMT      = '#,##0.00'
PCT_FMT        = '0.0"%"'


def _style_header_row(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def build_excel(user_id, year, enc_key, db) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    label = str(year) if year else "All Years"

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    cat_rows, grand_total = _get_category_summary(user_id, year, db)
    essential    = sum(float(r.total) for r in cat_rows
                       if r.category_name != 'Other')
    non_essential = grand_total - essential

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = f"fintrack — Financial Summary ({label})"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    ws.append([])
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append([])

    # KPIs
    for label_text, value in [
        ("Grand Total",      grand_total),
        ("Essential",        essential),
        ("Non-Essential",    non_essential),
    ]:
        row = ws.max_row + 1
        ws.append([label_text, value])
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).number_format = MONEY_FMT

    ws.append([])

    # Category table
    headers = ["Category", "Subcategory", "Amount", "% of Total", "Transactions"]
    ws.append(headers)
    _style_header_row(ws, ws.max_row, len(headers))

    for i, r in enumerate(cat_rows):
        pct = round(float(r.total) * 100 / grand_total, 1) if grand_total else 0
        ws.append([r.category_name, r.subcategory,
                   float(r.total), pct, r.txn_count])
        row_num = ws.max_row
        if i % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).fill = ALT_FILL
        ws.cell(row=row_num, column=3).number_format = MONEY_FMT
        ws.cell(row=row_num, column=4).number_format = '0.0"%"'
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = CELL_BORDER

    # Totals row
    ws.append(["TOTAL", "", grand_total, 100.0, sum(r.txn_count for r in cat_rows)])
    row_num = ws.max_row
    for col in range(1, 6):
        c = ws.cell(row=row_num, column=col)
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
        c.border = CELL_BORDER
    ws.cell(row=row_num, column=3).number_format = MONEY_FMT

    _auto_width(ws)

    # ── Sheet 2: Monthly Pivot ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Monthly Pivot")
    ws2.sheet_view.showGridLines = False

    pivot_rows = _get_monthly_pivot(user_id, year, db)

    # Build pivot structure
    months_seen = {}
    pivot = {}
    for row in pivot_rows:
        key   = (row.category_name, row.subcategory)
        month = row.month_label
        if month not in months_seen:
            months_seen[month] = (row.year_num, row.month_num)
        if key not in pivot:
            pivot[key] = {}
        pivot[key][month] = pivot[key].get(month, 0) + float(row.total)

    sorted_months = sorted(months_seen.keys(), key=lambda m: months_seen[m])

    # Header row
    header = ["Category", "Subcategory"] + sorted_months + ["Total"]
    ws2.append(header)
    _style_header_row(ws2, 1, len(header))

    col_totals = {m: 0.0 for m in sorted_months}
    grand = 0.0

    for i, ((cat, sub), month_data) in enumerate(sorted(pivot.items())):
        row_total = sum(month_data.values())
        grand += row_total
        row_vals = [cat, sub]
        for m in sorted_months:
            v = month_data.get(m, 0)
            col_totals[m] += v
            row_vals.append(v if v > 0 else None)
        row_vals.append(row_total)
        ws2.append(row_vals)
        row_num = ws2.max_row
        if i % 2 == 0:
            for col in range(1, len(header) + 1):
                ws2.cell(row=row_num, column=col).fill = ALT_FILL
        for col in range(3, len(header) + 1):
            ws2.cell(row=row_num, column=col).number_format = MONEY_FMT
        for col in range(1, len(header) + 1):
            ws2.cell(row=row_num, column=col).border = CELL_BORDER

    # Totals row
    totals = ["TOTAL", ""] + [col_totals[m] for m in sorted_months] + [grand]
    ws2.append(totals)
    row_num = ws2.max_row
    for col in range(1, len(header) + 1):
        c = ws2.cell(row=row_num, column=col)
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
        c.border = CELL_BORDER
        if col >= 3:
            c.number_format = MONEY_FMT

    _auto_width(ws2)

    # ── Sheet 3: Transactions ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Transactions")
    ws3.sheet_view.showGridLines = False

    txns = _get_transactions(user_id, year, enc_key, db)
    headers3 = ["Date", "Description", "Amount", "Category",
                "Subcategory", "Essential", "Card"]
    ws3.append(headers3)
    _style_header_row(ws3, 1, len(headers3))

    for i, t in enumerate(txns):
        ws3.append([
            t["date"], t["description"], t["amount"],
            t["category"], t["subcategory"], t["essential"], t["provider"]
        ])
        row_num = ws3.max_row
        if i % 2 == 0:
            for col in range(1, 8):
                ws3.cell(row=row_num, column=col).fill = ALT_FILL
        ws3.cell(row=row_num, column=3).number_format = MONEY_FMT
        for col in range(1, 8):
            ws3.cell(row=row_num, column=col).border = CELL_BORDER

    _auto_width(ws3)
    ws3.freeze_panes = "A2"

    # ── Sheet 4: Categories Detail ────────────────────────────────────────────
    ws4 = wb.create_sheet("Categories")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Category", "Subcategory", "Amount", "% of Total",
                "Transactions", "Avg per Transaction"]
    ws4.append(headers4)
    _style_header_row(ws4, 1, len(headers4))

    for i, r in enumerate(cat_rows):
        pct   = round(float(r.total) * 100 / grand_total, 1) if grand_total else 0
        avg   = round(float(r.total) / r.txn_count, 2) if r.txn_count else 0
        ws4.append([r.category_name, r.subcategory,
                    float(r.total), pct, r.txn_count, avg])
        row_num = ws4.max_row
        if i % 2 == 0:
            for col in range(1, 7):
                ws4.cell(row=row_num, column=col).fill = ALT_FILL
        for col in [3, 6]:
            ws4.cell(row=row_num, column=col).number_format = MONEY_FMT
        ws4.cell(row=row_num, column=4).number_format = '0.0"%"'
        for col in range(1, 7):
            ws4.cell(row=row_num, column=col).border = CELL_BORDER

    _auto_width(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── PDF Export ────────────────────────────────────────────────────────────────

def build_pdf(user_id, year, enc_key, db) -> bytes:
    buf = io.BytesIO()
    label = str(year) if year else "All Years"
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(letter),
        rightMargin=0.5*inch, leftMargin=0.5*inch,
        topMargin=0.5*inch, bottomMargin=0.5*inch
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=18, textColor=colors.HexColor("#1E3A8A"),
        spaceAfter=6
    )
    h2_style = ParagraphStyle(
        "H2", parent=styles["Heading2"],
        fontSize=13, textColor=colors.HexColor("#1565C0"),
        spaceBefore=16, spaceAfter=6
    )
    small = ParagraphStyle(
        "Small", parent=styles["Normal"],
        fontSize=8, textColor=colors.HexColor("#6B7280")
    )

    header_bg = colors.HexColor("#1565C0")
    alt_bg     = colors.HexColor("#EEF2FF")
    total_bg   = colors.HexColor("#DBEAFE")

    def money(v):
        return f"${float(v):,.2f}"

    def tbl_style(data, has_total=False):
        style = [
            ("BACKGROUND",  (0, 0), (-1, 0),  header_bg),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0),  9),
            ("ALIGN",       (0, 0), (-1, -1), "LEFT"),
            ("ALIGN",       (2, 1), (-1, -1), "RIGHT"),
            ("FONTSIZE",    (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2 if has_total else -1),
             [colors.white, alt_bg]),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        if has_total:
            style += [
                ("BACKGROUND", (0, -1), (-1, -1), total_bg),
                ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
            ]
        return TableStyle(style)

    story = []

    # Title
    story.append(Paragraph(f"fintrack — Financial Report ({label})", title_style))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        small
    ))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#1565C0")))
    story.append(Spacer(1, 0.15*inch))

    # ── Section 1: Summary KPIs ───────────────────────────────────────────────
    cat_rows, grand_total = _get_category_summary(user_id, year, db)
    essential     = sum(float(r.total) for r in cat_rows
                        if r.category_name != 'Other')
    non_essential = grand_total - essential

    story.append(Paragraph("Financial Summary", h2_style))
    kpi_data = [
        ["Metric", "Amount", "% of Total"],
        ["Grand Total",    money(grand_total), "100.0%"],
        ["Essential",      money(essential),
         f"{round(essential*100/grand_total,1)}%" if grand_total else "0%"],
        ["Non-Essential",  money(non_essential),
         f"{round(non_essential*100/grand_total,1)}%" if grand_total else "0%"],
        ["Transactions",   str(sum(r.txn_count for r in cat_rows)), ""],
    ]
    kpi_tbl = Table(kpi_data, colWidths=[3*inch, 2*inch, 1.5*inch])
    kpi_tbl.setStyle(tbl_style(kpi_data))
    story.append(kpi_tbl)
    story.append(Spacer(1, 0.2*inch))

    # ── Section 2: Category Summary ───────────────────────────────────────────
    story.append(Paragraph("Spending by Category", h2_style))
    cat_header = ["Category", "Subcategory", "Amount", "% of Total", "Txns"]
    cat_data = [cat_header]
    for r in cat_rows:
        pct = round(float(r.total)*100/grand_total, 1) if grand_total else 0
        cat_data.append([
            r.category_name, r.subcategory,
            money(r.total), f"{pct}%", str(r.txn_count)
        ])
    cat_data.append([
        "TOTAL", "", money(grand_total), "100.0%",
        str(sum(r.txn_count for r in cat_rows))
    ])
    col_w = [2.2*inch, 2.2*inch, 1.5*inch, 1.2*inch, 0.8*inch]
    cat_tbl = Table(cat_data, colWidths=col_w)
    cat_tbl.setStyle(tbl_style(cat_data, has_total=True))
    story.append(cat_tbl)
    story.append(Spacer(1, 0.2*inch))

    # ── Section 3: Monthly Pivot (top 10 categories) ──────────────────────────
    story.append(Paragraph("Monthly Spending Pivot", h2_style))
    pivot_rows = _get_monthly_pivot(user_id, year, db)

    months_seen = {}
    pivot = {}
    for row in pivot_rows:
        key   = (row.category_name, row.subcategory)
        month = row.month_label
        if month not in months_seen:
            months_seen[month] = (row.year_num, row.month_num)
        if key not in pivot:
            pivot[key] = {}
        pivot[key][month] = pivot[key].get(month, 0) + float(row.total)

    sorted_months = sorted(months_seen.keys(), key=lambda m: months_seen[m])
    short_months  = [m.split()[0] for m in sorted_months]

    pivot_header = ["Category"] + short_months + ["Total"]
    pivot_data   = [pivot_header]

    col_totals = {m: 0.0 for m in sorted_months}
    grand2 = 0.0

    # Top 15 categories by total
    sorted_cats = sorted(pivot.items(), key=lambda x: sum(x[1].values()), reverse=True)[:15]
    for (cat, sub), month_data in sorted_cats:
        row_total = sum(month_data.values())
        grand2 += row_total
        label_text = cat if cat == sub else f"{cat}/{sub[:8]}"
        row_vals = [label_text]
        for m in sorted_months:
            v = month_data.get(m, 0)
            col_totals[m] += v
            row_vals.append(f"${v:,.0f}" if v > 0 else "—")
        row_vals.append(f"${row_total:,.0f}")
        pivot_data.append(row_vals)

    pivot_data.append(
        ["TOTAL"] +
        [f"${col_totals[m]:,.0f}" for m in sorted_months] +
        [f"${grand2:,.0f}"]
    )

    n_cols = len(pivot_header)
    cat_col_w  = 1.6*inch
    month_col_w = (9.5*inch - cat_col_w) / max(n_cols - 1, 1)
    piv_widths = [cat_col_w] + [month_col_w] * (n_cols - 1)

    piv_tbl = Table(pivot_data, colWidths=piv_widths)
    piv_tbl.setStyle(tbl_style(pivot_data, has_total=True))
    story.append(piv_tbl)
    story.append(Spacer(1, 0.2*inch))

    # ── Section 4: Recent Transactions (last 50) ──────────────────────────────
    story.append(Paragraph("Recent Transactions (last 50)", h2_style))
    txns = _get_transactions(user_id, year, enc_key, db)[:50]
    txn_header = ["Date", "Description", "Amount", "Category", "Card"]
    txn_data   = [txn_header]
    for t in txns:
        desc = t["description"][:35] + "…" if len(t["description"]) > 35 else t["description"]
        txn_data.append([
            t["date"], desc, money(t["amount"]),
            t["category"], t["provider"]
        ])
    txn_tbl = Table(txn_data,
                    colWidths=[1.1*inch, 4.0*inch, 1.2*inch, 1.8*inch, 0.9*inch])
    txn_tbl.setStyle(tbl_style(txn_data))
    story.append(txn_tbl)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ── API Endpoints ─────────────────────────────────────────────────────────────

@router.get("/excel")
def export_excel(
    current_user: CurrentUser,
    year: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    password: str = Header(..., alias="x-fintrack-password"),
):
    """Download Excel workbook with 4 sheets: Summary, Monthly Pivot,
    Transactions, Categories."""
    enc_key  = _get_enc_key(current_user.id, password, db)
    filename = f"fintrack_{year or 'all'}.xlsx"
    data     = build_excel(current_user.id, year, enc_key, db)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/pdf")
def export_pdf(
    current_user: CurrentUser,
    year: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    password: str = Header(..., alias="x-fintrack-password"),
):
    """Download PDF report with summary, categories, pivot, and transactions."""
    enc_key  = _get_enc_key(current_user.id, password, db)
    filename = f"fintrack_{year or 'all'}.pdf"
    data     = build_pdf(current_user.id, year, enc_key, db)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
